"""
Lectura y conversión de archivos: PDF → imágenes, XLSX, TXT.
"""
import base64
import os
import re
import tempfile
from pathlib import Path
from typing import Optional
import openpyxl

# Extensiones soportadas
IMG_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.bmp', '.gif', '.webp', '.tiff', '.tif'}
PDF_EXTENSIONS = {'.pdf'}
EXCEL_EXTENSIONS = {'.xlsx', '.xls'}
TXT_EXTENSIONS = {'.txt'}

# Archivos a ignorar
IGNORE_PATTERNS = {'.tmp', '.lnk', '.dat', 'winmail', '.log',
                    '_extraccion.json', '_extraccion.txt', '_tablas.pdf',
                    '_tablas.json', '_tablas.txt', 'clasificacion.json'}


def should_ignore(path: Path) -> bool:
    name_lower = path.name.lower()
    for pat in IGNORE_PATTERNS:
        if pat in name_lower:
            return True
    return False


def image_to_base64(path: Path, max_bytes: int = 3_500_000) -> tuple[str, str]:
    """Devuelve (base64_data, media_type). Redimensiona/comprime si excede max_bytes.
    Límite API Claude es 5MB en BASE64; base64 infla ~33%, así que un archivo
    de 3.5MB binario ≈ 4.7MB base64. Margen seguro."""
    ext = path.suffix.lower()
    media_map = {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
        '.png': 'image/png', '.gif': 'image/gif',
        '.webp': 'image/webp', '.bmp': 'image/png',
    }
    media_type = media_map.get(ext, 'image/jpeg')

    size = path.stat().st_size
    if size <= max_bytes:
        with open(path, 'rb') as f:
            data = base64.standard_b64encode(f.read()).decode('utf-8')
        return data, media_type

    # Imagen grande: reescalar y guardar a JPEG con calidad progresiva
    from PIL import Image
    import io
    img = Image.open(path)
    img = img.convert("RGB")
    # Arrancar con escala basada en ratio de tamaño
    scale = (max_bytes / size) ** 0.5
    w = max(1, int(img.width * scale))
    h = max(1, int(img.height * scale))
    img_small = img.resize((w, h), Image.LANCZOS)
    for quality in (85, 75, 65, 55):
        buf = io.BytesIO()
        img_small.save(buf, format="JPEG", quality=quality, optimize=True)
        if buf.tell() <= max_bytes:
            data = base64.standard_b64encode(buf.getvalue()).decode('utf-8')
            return data, "image/jpeg"
    # Último recurso: reducir más
    img_small = img_small.resize((w // 2, h // 2), Image.LANCZOS)
    buf = io.BytesIO()
    img_small.save(buf, format="JPEG", quality=60, optimize=True)
    data = base64.standard_b64encode(buf.getvalue()).decode('utf-8')
    return data, "image/jpeg"


_easyocr_reader = None

def _get_easyocr_reader():
    """Inicializa EasyOCR una sola vez (carga el modelo ~1-2s)."""
    global _easyocr_reader
    if _easyocr_reader is None:
        import easyocr
        _easyocr_reader = easyocr.Reader(['es', 'en'], verbose=False)
    return _easyocr_reader


def run_easyocr_on_image(img) -> str:
    """
    Ejecuta EasyOCR sobre una imagen PIL y devuelve el texto reconocido
    como string, con cada detección en una línea junto a su confianza.
    """
    try:
        import numpy as np
        reader = _get_easyocr_reader()
        img_np = np.array(img.convert('RGB'))
        results = reader.readtext(img_np, detail=1, paragraph=False)
        lines = []
        for (_, text, conf) in results:
            if conf > 0.1:  # filtrar detecciones muy inciertas
                lines.append(f"{text} [{conf:.0%}]")
        return '\n'.join(lines)
    except Exception as e:
        return f"[EasyOCR error: {e}]"


def pdf_extract_text(path: Path, max_pages: int = 5, min_chars_per_page: int = 80) -> Optional[str]:
    """
    Intenta extraer texto de un PDF con capa de texto (PDFs digitales).
    Devuelve el texto extraído si el PDF tiene contenido textual suficiente,
    o None si el PDF es una imagen escaneada (sin texto).

    El umbral min_chars_per_page evita falsos positivos en PDFs con poco texto
    (ej: solo cabeceras) que en realidad necesitan visión.
    """
    try:
        import pdfplumber
        with pdfplumber.open(str(path)) as pdf:
            pages = pdf.pages[:max_pages]
            all_text = []
            total_chars = 0
            for page in pages:
                text = page.extract_text() or ''
                all_text.append(text)
                total_chars += len(text)
            # Verificar que hay texto suficiente (no es un escaneado)
            avg_chars = total_chars / max(len(pages), 1)
            if avg_chars < min_chars_per_page:
                return None
            return '\n\n--- PÁGINA SIGUIENTE ---\n\n'.join(t for t in all_text if t)
    except Exception:
        return None


def pdf_pages_to_base64(path: Path, dpi: int = 200, max_pages: int = 5, return_pil: bool = False) -> list:
    """
    Convierte páginas de un PDF a lista de (base64, media_type).
    Si return_pil=True, devuelve (base64, media_type, pil_image) para usar con EasyOCR.
    """
    try:
        from pdf2image import convert_from_path
        images = convert_from_path(str(path), dpi=dpi, first_page=1, last_page=max_pages)
        result = []
        for img in images:
            with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as tmp:
                img.save(tmp.name, 'JPEG', quality=85)
                tmp_path = Path(tmp.name)
            # Verificar tamaño — API de Claude limita a 5MB por imagen
            size = tmp_path.stat().st_size
            if size > 4_500_000:
                # Reescalar al 75% y recomprimir con menor calidad
                img_small = img.resize(
                    (int(img.width * 0.75), int(img.height * 0.75))
                )
                img_small.save(tmp.name, 'JPEG', quality=75)
                size2 = tmp_path.stat().st_size
                if size2 > 4_500_000:
                    # Segundo intento: 50% del original
                    img_small2 = img.resize(
                        (int(img.width * 0.5), int(img.height * 0.5))
                    )
                    img_small2.save(tmp.name, 'JPEG', quality=70)
            with open(tmp_path, 'rb') as f:
                data = base64.standard_b64encode(f.read()).decode('utf-8')
            tmp_path.unlink(missing_ok=True)
            if return_pil:
                result.append((data, 'image/jpeg', img))
            else:
                result.append((data, 'image/jpeg'))
        return result
    except Exception as e:
        print(f"  [!] Error convirtiendo PDF {path.name}: {e}")
        return []


def read_excel_as_text(path: Path) -> str:
    """Lee un XLSX y devuelve su contenido como texto estructurado."""
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"=== HOJA: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                if any(c is not None for c in row):
                    # Filtrar filas completamente vacías o de errores
                    vals = [str(c) if c is not None else '' for c in row]
                    # Solo incluir si hay algo útil
                    non_empty = [v for v in vals if v.strip() and v != '#N/A' and v != 'None']
                    if non_empty:
                        lines.append('\t'.join(vals).rstrip())
        return '\n'.join(lines)
    except Exception as e:
        return f"[Error leyendo Excel: {e}]"


def read_txt(path: Path) -> str:
    """Lee un TXT con detección de encoding (incluye BOM e iso-8859-15,
    habituales en archivos del taller)."""
    for enc in ['utf-8', 'utf-8-sig', 'iso-8859-15', 'latin-1', 'cp1252']:
        try:
            return path.read_text(encoding=enc)
        except Exception:
            continue
    return f"[Error leyendo TXT {path.name}: ningún encoding estándar funcionó]"


def _score_pdf(path: Path) -> int:
    """
    Puntúa un PDF para priorizar cuáles enviar a Claude.
    Mayor puntuación = más importante.
    """
    name = path.name.lower()
    score = 0
    # Plantillas de presupuesto marmolista — MÁS IMPORTANTES
    if 'plantilla' in name or 'presupuesto' in name:
        score += 100
    # Planos del programa de diseño
    if 'planta' in name or 'encimera' in name or 'diseño' in name or 'cocina' in name:
        score += 80
    # Presupuestos del marmolista — ordenar por número (mayor = más reciente)
    if name.startswith('pr') or name.startswith('f'):
        # Extraer número del presupuesto para priorizar el más reciente
        nums = re.findall(r'\d{4}', name)
        if nums:
            score += 50 + int(nums[0]) // 100  # más reciente = número más alto
    # PDFs escaneados (fechas en nombre)
    if re.search(r'\d{14}', name):
        score += 60
    return score


def _es_mgr_marmolista(path: Path) -> bool:
    """
    Detecta archivos del MGR/presupuesto del marmolista por heurística:
    - PDFs cuyo nombre empieza por 'PR\\d+'  (ej: PR1382_J0015...)
    - PDFs cuyo nombre empieza por 'F\\d+'   (facturas: F250121_...)
    - Excels cuyo nombre coincide con el código del trabajo + descripción
      (suelen ser exports del MGR, contienen 'Presupuesto' interno).

    El extractor IGNORA estos archivos: solo trabaja con plano, plantilla,
    fotos y anotaciones del operador.
    """
    name = path.name
    # PDFs de marmolista: PR\d+ o F\d+
    if path.suffix.lower() == '.pdf':
        if re.match(r'^(PR|F)\d+', name):
            return True
    # Excels: si la primera hoja se llama 'Presupuesto NNNN', es del marmolista
    if path.suffix.lower() == '.xlsx':
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=False)
            sheet_names_lower = [s.lower() for s in wb.sheetnames]
            wb.close()
            for s in sheet_names_lower:
                if 'presupuesto' in s:
                    return True
            # También detectar por contenido típico (col 1 fila 14 = 'Ref.')
        except Exception:
            pass
    return False


def collect_files(folder: Path, max_pdfs: int = 5,
                   excluir_mgr: bool = True) -> dict:
    """
    Recopila y prioriza los archivos de una carpeta de trabajo.
    Solo lee archivos de la carpeta raíz; las subcarpetas (Segundas, Terceras,
    Cuartas, Quintas, etc.) corresponden a revisiones posteriores y NO se
    procesan en este flujo.

    `excluir_mgr=True` (default): omite los archivos del presupuesto del
    marmolista (PR*.pdf, F*.pdf, *.xlsx con hojas 'Presupuesto'). El
    extractor solo usa plano + plantilla + anotaciones — la forma de la
    encimera viene SOLO de esas fuentes.
    """
    files = {
        'images': [],
        'pdfs': [],
        'excels': [],
        'txts': [],
        'ignored': [],
        'subfolders': [],
        'pdfs_omitidos': [],
        'mgr_excluidos': [],   # archivos MGR ignorados (informativo)
    }

    all_pdfs = []

    for f in sorted(folder.iterdir()):
        if f.is_dir():
            if not should_ignore(f):
                files['subfolders'].append(f)
            continue
        if not f.is_file() or should_ignore(f):
            continue
        if excluir_mgr and _es_mgr_marmolista(f):
            files['mgr_excluidos'].append(f)
            continue
        ext = f.suffix.lower()
        if ext in IMG_EXTENSIONS:
            files['images'].append(f)
        elif ext in PDF_EXTENSIONS:
            all_pdfs.append((f, 'Primeras'))
        elif ext in EXCEL_EXTENSIONS:
            files['excels'].append(f)
        elif ext in TXT_EXTENSIONS:
            files['txts'].append(f)

    all_pdfs.sort(key=lambda x: _score_pdf(x[0]), reverse=True)
    files['pdfs'] = all_pdfs[:max_pdfs]
    files['pdfs_omitidos'] = all_pdfs[max_pdfs:]

    return files


def build_claude_content(folder: Path, verbose: bool = True, max_pdfs: int = 5,
                          clasificacion: Optional[dict] = None) -> tuple[list, list[str]]:
    """
    Construye el contenido para enviar a Claude (lista de bloques de contenido).
    Devuelve (content_blocks, archivos_procesados).

    Si se pasa `clasificacion` (output de clasificador.py), cada imagen/página
    lleva una etiqueta [TIPO: <categoria>] para que Claude sepa qué tipo de
    contenido tiene y pueda enrutarlo correctamente:
      - forma → solo plano-*
      - cantidades/precios → MGR-tabla
      - huecos/material → plantilla-texto
      - perspectiva-3D → solo contexto, NO cotas

    Si existe `anotaciones.json` en la carpeta del proyecto:
      - Las páginas anotadas se sustituyen por su imagen anotada (con
        overlay de rotulador del operador)
      - Se inserta un bloque de notas globales y leyenda de colores
        antes del contenido
    """
    files = collect_files(folder, max_pdfs=max_pdfs)
    content = []
    archivos = []
    clasif = clasificacion.get("archivos", {}) if clasificacion else {}

    # Cargar anotaciones manuales si existen
    anot_path = folder / "anotaciones.json"
    anotaciones = {}
    paginas_anot_dir = folder / "anotaciones"
    if anot_path.exists():
        try:
            import json as _json
            anotaciones = _json.loads(anot_path.read_text(encoding="utf-8"))
            if not isinstance(anotaciones.get("paginas_anotadas"), dict):
                print(f"  ⚠ anotaciones.json con esquema inesperado en "
                      f"{folder.name} — los trazos del operador NO se usarán")
        except Exception as e:
            print(f"  ⚠ anotaciones.json corrupto en {folder.name}: {e} — "
                  f"los trazos del operador NO se usarán")
            anotaciones = {}

    def _safe_pagina_filename(pid: str) -> str:
        return re.sub(r"[^A-Za-z0-9._-]", "_", pid) + ".png"

    def _imagen_anotada_b64(pagina_id: str) -> Optional[tuple[str, str]]:
        """Composite (original + overlay del operador). Útil para context."""
        info = (anotaciones.get("paginas_anotadas") or {}).get(pagina_id)
        if not info or not info.get("tiene_trazos"):
            return None
        fname = _safe_pagina_filename(pagina_id)
        path = paginas_anot_dir / fname
        if not path.exists():
            return None
        try:
            data, mt = image_to_base64(path)
            return data, mt
        except Exception:
            return None

    def _overlay_solo_b64(pagina_id: str) -> Optional[tuple[str, str]]:
        """Overlay aislado del operador con fondo blanco para máxima claridad
        de QUÉ marcó. Las zonas no pintadas son blancas (no transparentes)."""
        info = (anotaciones.get("paginas_anotadas") or {}).get(pagina_id)
        if not info or not info.get("tiene_trazos"):
            return None
        fname = _safe_pagina_filename(pagina_id)
        overlay_path = paginas_anot_dir / ("overlay_" + fname)
        if not overlay_path.exists():
            return None
        try:
            from PIL import Image
            import io as _io
            ov = Image.open(overlay_path).convert("RGBA")
            # Componer sobre fondo blanco para que los trazos contrasten al máximo
            blanco = Image.new("RGBA", ov.size, (255, 255, 255, 255))
            sobre_blanco = Image.alpha_composite(blanco, ov).convert("RGB")
            buf = _io.BytesIO()
            sobre_blanco.save(buf, format="PNG", optimize=True)
            data = base64.standard_b64encode(buf.getvalue()).decode("utf-8")
            return data, "image/png"
        except Exception:
            return None

    def _tag(filename: str, pagina: Optional[int] = None) -> str:
        """Devuelve etiqueta [TIPO: ...] si tenemos clasificación del archivo/página."""
        info = clasif.get(filename)
        if not info or "error" in info:
            return ""
        if pagina is not None and "paginas" in info:
            for p in info["paginas"]:
                if p.get("pagina") == pagina:
                    cat = p.get("categoria", "?")
                    cotas = ", ".join(str(c) for c in (p.get("cotas_visibles_mm") or [])[:8])
                    cotas_s = f" cotas:[{cotas}]" if cotas else ""
                    return f"[TIPO:{cat}{cotas_s}]"
        elif "categoria" in info:
            cat = info.get("categoria", "?")
            return f"[TIPO:{cat}]"
        return ""

    segundas = [p for p, lbl in files['pdfs'] if lbl != 'Primeras']
    if verbose:
        omit_info = f", PDFs omitidos: {len(files['pdfs_omitidos'])}" if files['pdfs_omitidos'] else ""
        sub_info = f", Subcarpetas: {[s.name for s in files['subfolders']]}" if files['subfolders'] else ""
        seg_info = f", PDFs revisados (Segundas/Terceras): {len(segundas)}" if segundas else ""
        clasif_info = f", clasificado: ✓" if clasif else ""
        mgr_info = f", MGR-excluidos: {len(files.get('mgr_excluidos',[]))}" if files.get('mgr_excluidos') else ""
        print(f"  Imágenes: {len(files['images'])}, PDFs: {len(files['pdfs'])}, "
              f"Excels: {len(files['excels'])}, TXTs: {len(files['txts'])}"
              f"{omit_info}{sub_info}{seg_info}{clasif_info}{mgr_info}")

    # 1. Texto inicial de contexto
    ctx = f"Carpeta de trabajo: {folder.name}\n\nA continuación tienes los archivos del trabajo:"
    if clasif:
        ctx += ("\n\n📋 CLASIFICACIÓN DE FUENTES (úsala para enrutar tu atención):\n"
                "  • Cada imagen/página viene con una etiqueta [TIPO:xxx].\n"
                "  • Para FORMA y geometría de piezas: usa SOLO 'plano-planta-2D-cad', 'plano-anotado' y 'plano-manuscrito'.\n"
                "  • Para PRECIOS, CANTIDADES TOTALES, m² facturados: usa SOLO 'MGR-tabla'.\n"
                "  • Para HUECOS, MATERIAL, MARCA, COLOR, GROSOR: usa 'plantilla-texto'.\n"
                "  • 'perspectiva-3D' y 'foto-cocina': solo contexto. NO tomes cotas de ahí.\n"
                "  • 'whatsapp-plantilla': trátala como plantilla manuscrita o foto según calidad.\n"
                "  • 'otro': ignorar.\n"
                "  • Si extraes una cota de la imagen, indica de qué [TIPO] vino.")
    if files['pdfs_omitidos']:
        ctx += f"\n\nNOTA: Por límite de contexto se han omitido {len(files['pdfs_omitidos'])} PDFs menos prioritarios: "
        ctx += ", ".join(f.name for f, _ in files['pdfs_omitidos'])
    if files['subfolders']:
        ctx += f"\n\nNOTA: Esta carpeta tiene subcarpetas con versiones adicionales: "
        ctx += ", ".join(s.name for s in files['subfolders'])
        ctx += ". Procesa los archivos de la carpeta raíz como la versión más definitiva."

    # Anotaciones manuales: notas globales + leyenda de colores
    if anotaciones:
        notas = anotaciones.get("notas_globales") or []
        n_anotadas = sum(1 for v in (anotaciones.get("paginas_anotadas") or {}).values()
                         if v.get("tiene_trazos"))
        if notas or n_anotadas:
            ctx += ("\n\n🖍 ANOTACIONES MANUALES DEL OPERADOR — AUTORIDAD ABSOLUTA SOBRE LA FORMA "
                    "(prevalece al MGR, al plano sin marcar y a TODA otra fuente):")
            if n_anotadas:
                ctx += (f"\nHay {n_anotadas} página(s) con TRAZOS DE ROTULADOR del operador "
                        f"sobre el plano original. Las verás como áreas coloreadas "
                        f"semi-transparentes encima de la imagen. Significa:\n"
                        f"  🟡 amarillo  = encimera/isla\n"
                        f"  🔵 azul      = frontal/chapeado\n"
                        f"  🟣 lila      = zócalo/rodapié\n"
                        f"  🟠 naranja   = copete\n"
                        f"  🟢 verde     = costado/cascada\n"
                        f"  🟫 marrón    = pilar\n"
                        f"  🔴 rojo      = hueco (placa/fregadero/grifo/enchufe)\n"
                        f"\n⚠ JERARQUÍA DE FUENTES (estricta):\n"
                        f"  1) Anotaciones manuales del operador (color + cotas redondeadas)\n"
                        f"  2) Cotas explícitas escritas en el plano cerca de un trazo\n"
                        f"  3) Plano sin marcar (cotas generales)\n"
                        f"  4) MGR / Excel / presupuesto (referencia secundaria, BOUNDING)\n\n"
                        f"⚠ EL MGR ES BOUNDING / MERMA, NO FORMA REAL. Si el operador marca\n"
                        f"  encimera amarilla con cota '2700' y el MGR dice '3300', LA REALIDAD ES 2700.\n"
                        f"  Los 600mm de diferencia son zona oculta, merma de pilar, o cobranza\n"
                        f"  por área no fabricada. NO inventes piezas que el MGR sugiere pero el\n"
                        f"  operador NO marcó.\n\n"
                        f"⚠ MARRÓN = PILAR (saliente del muro hacia la encimera): genera 1-3 piezas\n"
                        f"  PEQUEÑAS alrededor del pilar (lados ~150-300mm × fondo + frente ~600mm).\n"
                        f"  NO confundir con un tramo de encimera grande de 1m+. Si el MGR dice\n"
                        f"  un tramo lateral 1020×350 pero el operador marcó pilar marrón ahí,\n"
                        f"  son las piezas de wrap del pilar, NO una encimera lateral.\n\n"
                        f"⚠ CADA ÁREA AMARILLA distinta = 1 PIEZA de encimera con sus propias cotas.\n"
                        f"  Si solo hay 1 área amarilla continua → 1 sola encimera.\n"
                        f"  Si hay 2 áreas amarillas → 2 encimeras.\n"
                        f"  NO añadas encimeras 'fantasma' que el operador no marcó.\n\n"
                        f"⚠ REGLA INVERSA — DEFAULT NO-ENCIMERA: Si una zona del plano NO tiene\n"
                        f"  trazo amarillo del operador, NO HAY encimera ahí, NO IMPORTA qué diga\n"
                        f"  el MGR o las tablas de nesting previas. Aunque el MGR liste un\n"
                        f"  'tramo lateral 1020×350', si esa zona está marcada con marrón (pilar),\n"
                        f"  son las piezas de wrap del pilar, NO una encimera. La diferencia entre\n"
                        f"  el área amarilla marcada y el bounding del MGR es MERMA cobrada\n"
                        f"  (zona detrás de pilar, vuelo, cobranza por área no fabricada).\n\n"
                        f"⚠ CUENTA LAS PIEZAS DE ENCIMERA: emitirás EXACTAMENTE el número de áreas\n"
                        f"  amarillas distintas marcadas. Si el operador marcó 1 sola área amarilla,\n"
                        f"  emites 1 encimera. Si marcó 2, emites 2. Sin excepciones.\n")
            if notas:
                ctx += "\n\nNotas globales del proyecto (texto del operador):"
                for n in notas:
                    ctx += f"\n  • {n}"
                ctx += ("\nEstas notas son AUTORITATIVAS. Si dicen 'cascada al suelo' emite una "
                        "pieza costado de la altura indicada. Si dicen 'pilar atrás' emite las "
                        "piezas alrededor del pilar. No las ignores.")

            # ── POLILÍNEAS EXACTAS (vértices crudos del operador) ────────
            polilineas_text = []
            for pid, pdata in (anotaciones.get("paginas_anotadas") or {}).items():
                cw = pdata.get("canvas_w")
                ch = pdata.get("canvas_h")
                polys_aqui = []
                for et in pdata.get("etiquetas", []):
                    if et.get("modo") == "poly" and et.get("puntos"):
                        polys_aqui.append({
                            "tipo": et.get("tipo"),
                            "color": et.get("color"),
                            "cerrado": et.get("cerrado", True),
                            "puntos": et["puntos"],
                        })
                if polys_aqui:
                    polilineas_text.append({"pagina": pid, "canvas": [cw, ch], "polilineas": polys_aqui})

            # ── PAREDES, MUEBLES ALTOS Y PILARES (ground truth del operador) ──
            paredes_por_pagina = []
            ma_por_pagina = []
            pilares_por_pagina = []
            for pid, pdata in (anotaciones.get("paginas_anotadas") or {}).items():
                cw = pdata.get("canvas_w")
                ch = pdata.get("canvas_h")
                paredes_aqui = []
                ma_aqui = []
                pilares_aqui = []
                for et in pdata.get("etiquetas", []):
                    pts = et.get("puntos")
                    if not pts:
                        continue
                    tipo = et.get("tipo")
                    if tipo == "pared":
                        paredes_aqui.append({"modo": et.get("modo"), "puntos": pts})
                    elif tipo == "muebles_altos":
                        ma_aqui.append({"modo": et.get("modo"), "puntos": pts})
                    elif tipo == "pilar":
                        pilares_aqui.append({"modo": et.get("modo"), "puntos": pts})
                if paredes_aqui:
                    paredes_por_pagina.append({"pagina": pid, "canvas": [cw, ch], "trazos": paredes_aqui})
                if ma_aqui:
                    ma_por_pagina.append({"pagina": pid, "canvas": [cw, ch], "trazos": ma_aqui})
                if pilares_aqui:
                    pilares_por_pagina.append({"pagina": pid, "canvas": [cw, ch], "trazos": pilares_aqui})

            if pilares_por_pagina:
                ctx += ("\n\n🏛️ PILARES — OBSTÁCULOS DEL EDIFICIO:\n"
                        "El operador ha marcado pilares estructurales (marrón). Un pilar NO "
                        "es una pieza a fabricar — es un OBSTÁCULO en el plano. La encimera "
                        "debe tener un HUECO o una MUESCA (forma L/U) que rodee el pilar para "
                        "esquivarlo en la instalación.\n\n"
                        "ACCIÓN OBLIGATORIA: Cuando hay un pilar marcado, los `vertices_mm` "
                        "de la encimera tienen que reflejar la muesca alrededor del pilar (no "
                        "puede ser un rectángulo plano que solape el pilar). Detalla en `notas` "
                        "que la muesca esquiva un pilar de las dimensiones aproximadas leídas.\n"
                        "PROHIBIDO: emitir piezas tipo 'pilar' o 'pilastra'. El pilar NUNCA es "
                        "una pieza de piedra, sólo afecta la geometría de la encimera.\n")
                for p in pilares_por_pagina:
                    ctx += f"\nPilar en página '{p['pagina']}' (canvas {p['canvas'][0]}×{p['canvas'][1]}px):\n"
                    for tr in p["trazos"]:
                        n = len(tr["puntos"])
                        muestra = tr["puntos"][:6] + (["..."] if n > 6 else [])
                        ctx += f"  - {tr['modo']} ({n}pt): {muestra}\n"

            if paredes_por_pagina or ma_por_pagina:
                ctx += ("\n\n🧱 PAREDES Y MUEBLES ALTOS — GROUND TRUTH DEL OPERADOR:\n"
                        "El operador ha trazado las paredes (gris oscuro) y/o las proyecciones "
                        "de los muebles altos (gris claro) sobre el plano. Son **autoritativos**.\n\n"
                        "REGLA DE PULIDOS: Una arista de la encimera que esté pegada a un trazo "
                        "de PARED o de MUEBLES_ALTOS NO se pule. Solo se pulen las aristas que dan "
                        "a espacio libre de la cocina. Si un trazo de pared cae sobre/junto a una "
                        "arista del polígono de la encimera, esa arista NO emite `recto_pulido` ni "
                        "participa en ingletes.\n\n"
                        "REGLA DE FRONTAL/CHAPEADO: La pieza FRONTAL (chapeado) sólo va donde haya "
                        "muebles bajos (espacio entre encimera y suelo) que reciban revestimiento — "
                        "típicamente bajo la encimera contra una pared. La presencia de MUEBLES_ALTOS "
                        "en la zona NO implica frontal allí.\n")
                for p in paredes_por_pagina:
                    ctx += f"\nPared en página '{p['pagina']}' (canvas {p['canvas'][0]}×{p['canvas'][1]}px):\n"
                    for tr in p["trazos"]:
                        n = len(tr["puntos"])
                        muestra = tr["puntos"][:6] + (["..."] if n > 6 else [])
                        ctx += f"  - {tr['modo']} ({n}pt): {muestra}\n"
                for p in ma_por_pagina:
                    ctx += f"\nMuebles altos en página '{p['pagina']}' (canvas {p['canvas'][0]}×{p['canvas'][1]}px):\n"
                    for tr in p["trazos"]:
                        n = len(tr["puntos"])
                        muestra = tr["puntos"][:6] + (["..."] if n > 6 else [])
                        ctx += f"  - {tr['modo']} ({n}pt): {muestra}\n"

            if polilineas_text:
                ctx += ("\n\n📐 POLILÍNEAS EXACTAS DEL OPERADOR — VÉRTICES AUTORITATIVOS:\n"
                        "El operador trazó polilíneas con click vértice a vértice. "
                        "Estos vértices son la GEOMETRÍA EXACTA de las piezas, en pixel-space del canvas.\n"
                        "Para emitir `vertices_mm` de cada pieza:\n"
                        "  1. Identifica las cotas reales del plano (ej: 4750mm de largo, 2635mm de ancho).\n"
                        "  2. Calcula UN factor de escala uniforme mm/pixel (mismo en X y Y) comparando "
                        "los lados de la polilínea con cotas conocidas.\n"
                        "  3. Aplica `vertices_mm[i] = (px[i] * escala, py[i] * escala)` a TODOS los puntos. "
                        "**NO INVIERTAS ningún eje**: si el pixel x crece hacia la derecha, mm x crece hacia la "
                        "derecha; si pixel y crece hacia abajo, mm y crece hacia abajo. NO espejes X. NO inviertas Y. "
                        "El programa de visualización CAD aplicará la inversión Y para CAD por su cuenta.\n"
                        "  4. NO añadas ni quites vértices — el número de vértices del polígono real es "
                        "EXACTAMENTE el número de puntos de la polilínea.\n"
                        "  5. NO reordenes los puntos — emítelos en el MISMO orden recibido.\n"
                        "  6. Trasláda al origen restando el min(x), min(y) de los vértices escalados, para "
                        "que el polígono quede en el cuadrante positivo `[0, largo] × [0, ancho]`.\n")
                for p in polilineas_text:
                    ctx += f"\nPágina '{p['pagina']}' (canvas {p['canvas'][0]}×{p['canvas'][1]}px):\n"
                    for poly in p["polilineas"]:
                        cerr = "cerrada" if poly["cerrado"] else "abierta"
                        ctx += f"  - {poly['tipo']} ({cerr}, {len(poly['puntos'])}v): {poly['puntos']}\n"

    content.append({"type": "text", "text": ctx})

    # 2. TXTs
    for txt_path in files['txts']:
        texto = read_txt(txt_path)
        content.append({
            "type": "text",
            "text": f"\n--- ARCHIVO TXT: {txt_path.name} ---\n{texto}"
        })
        archivos.append(txt_path.name)

    # 3. Excels
    for xl_path in files['excels']:
        texto = read_excel_as_text(xl_path)
        content.append({
            "type": "text",
            "text": f"\n--- ARCHIVO EXCEL: {xl_path.name} ---\n{texto}"
        })
        archivos.append(xl_path.name)

    # 4. PDFs → texto (si digital) o imágenes (si escaneado)
    for pdf_path, pdf_label in files['pdfs']:
        label_str = f" [{pdf_label}]" if pdf_label != 'Primeras' else ""
        label_header = f" — MEDIDAS REVISADAS ({pdf_label})" if pdf_label != 'Primeras' else ""
        name_lower = pdf_path.name.lower()

        # Plantillas manuscritas: siempre usar imágenes (escritura a mano, no digital)
        is_handwritten = any(w in name_lower for w in ('plantilla', 'presupuesto encimera', 'encimera '))

        # Intentar extracción de texto para PDFs no manuscritos
        pdf_text = None
        if not is_handwritten:
            pdf_text = pdf_extract_text(pdf_path)

        if pdf_text:
            # PDF digital con texto — enviar como texto (mucho más barato)
            if verbose:
                print(f"  Texto PDF: {pdf_path.name}{label_str} [{len(pdf_text)} chars]")
            content.append({
                "type": "text",
                "text": f"\n--- PDF (TEXTO): {pdf_path.name}{label_header} ---\n{pdf_text}"
            })
            archivos.append(pdf_path.name)
        else:
            # PDF escaneado o con elementos gráficos — usar imágenes
            if verbose:
                print(f"  Convirtiendo PDF: {pdf_path.name}{label_str}")
            dpi = 250 if is_handwritten else 200
            pages = pdf_pages_to_base64(pdf_path, dpi=dpi, return_pil=is_handwritten)
            if pages:
                content.append({
                    "type": "text",
                    "text": f"\n--- PDF: {pdf_path.name} ({len(pages)} páginas){label_header} ---"
                })
                for i, page_data in enumerate(pages):
                    if is_handwritten and isinstance(page_data, tuple) and len(page_data) == 3:
                        data, media_type, pil_img = page_data
                        # Pre-OCR con EasyOCR para ayudar a Claude con la letra manuscrita
                        if verbose:
                            print(f"    EasyOCR pág {i+1}...")
                        ocr_text = run_easyocr_on_image(pil_img)
                        if ocr_text:
                            content.append({
                                "type": "text",
                                "text": f"[EasyOCR pág {i+1} — texto detectado, úsalo como pista para leer la letra manuscrita]:\n{ocr_text}"
                            })
                    else:
                        data, media_type = page_data[0], page_data[1]
                    pagina_id = f"pdf::{pdf_path.name}::{i+1}"
                    overlay_solo = _overlay_solo_b64(pagina_id)
                    tag = _tag(pdf_path.name, pagina=i+1)
                    if overlay_solo:
                        # Tiene anotación: enviar 2 imágenes — original + overlay solo
                        content.append({"type": "text",
                                         "text": f"[Página {i+1} de '{pdf_path.name}' — IMAGEN ORIGINAL del plano] {tag}"})
                        content.append({"type": "image", "source": {
                            "type": "base64", "media_type": media_type, "data": data}})
                        content.append({"type": "text",
                                         "text": (f"[Página {i+1} de '{pdf_path.name}' — OVERLAY DEL OPERADOR sobre fondo blanco]\n"
                                                  f"Esta imagen muestra ÚNICAMENTE los trazos del rotulador del operador. "
                                                  f"Las áreas blancas NO ESTÁN MARCADAS y por tanto NO existen como pieza física. "
                                                  f"Solo emite piezas en zonas con color en esta imagen. Usa la imagen original "
                                                  f"de arriba para leer las cotas exactas (mm).")})
                        ov_data, ov_mt = overlay_solo
                        content.append({"type": "image", "source": {
                            "type": "base64", "media_type": ov_mt, "data": ov_data}})
                    else:
                        if tag:
                            content.append({"type": "text",
                                             "text": f"[Página {i+1} de '{pdf_path.name}'] {tag}"})
                        content.append({"type": "image", "source": {
                            "type": "base64", "media_type": media_type, "data": data}})
                archivos.append(pdf_path.name)

    # 5. Imágenes directas
    for img_path in files['images']:
        try:
            data, media_type = image_to_base64(img_path)
            pagina_id = f"img::{img_path.name}"
            overlay_solo = _overlay_solo_b64(pagina_id)
            tag = _tag(img_path.name)
            if overlay_solo:
                content.append({"type": "text",
                                 "text": f"\n--- IMAGEN ORIGINAL: {img_path.name} --- {tag}"})
                content.append({"type": "image", "source": {
                    "type": "base64", "media_type": media_type, "data": data}})
                content.append({"type": "text",
                                 "text": (f"--- OVERLAY DEL OPERADOR ({img_path.name}) sobre fondo blanco ---\n"
                                          f"Trazos de rotulador del operador SÓLOS. Áreas blancas = no marcadas = no hay pieza ahí. "
                                          f"Cotas reales: léelas de la imagen original de arriba.")})
                ov_data, ov_mt = overlay_solo
                content.append({"type": "image", "source": {
                    "type": "base64", "media_type": ov_mt, "data": ov_data}})
                archivos.append(img_path.name)
                continue
            content.append({
                "type": "text",
                "text": f"\n--- IMAGEN: {img_path.name} --- {tag}"
            })
            content.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": media_type,
                    "data": data,
                }
            })
            archivos.append(img_path.name)
        except Exception as e:
            print(f"  [!] Error con imagen {img_path.name}: {e}")

    return content, archivos
