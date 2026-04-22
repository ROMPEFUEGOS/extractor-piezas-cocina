#!/usr/bin/env python3
"""
verificar_contra_excel.py — Compara una extracción JSON contra el Excel MGR del
mismo proyecto (ground truth) y reporta discrepancias para mejorar el extractor.

Flujo:
    1. Lee el Excel MGR (última hoja "Presupuesto NNNN" por fecha)
    2. Parsea las líneas relevantes (material, huecos, m², ml cantos)
    3. Compara con el JSON de extracción del mismo proyecto
    4. Imprime un reporte con ✓ (coincide) y ✗ (difiere)

Uso:
    python3 verificar_contra_excel.py "/ruta/J0221_.../J0221_extraccion.json"
    python3 verificar_contra_excel.py "/ruta/J0221_.../"
    python3 verificar_contra_excel.py --lote "/ruta/Cocimoble2025/"   # todos los JSONs
    python3 verificar_contra_excel.py --lote --csv resumen.csv

Salida:
  Por proyecto: lista de (concepto, valor_excel, valor_json, match/diff)
  Lote: CSV agregado + contadores de discrepancias típicas
"""

import argparse
import csv
import json
import re
import sys
from pathlib import Path
from typing import Optional
from datetime import datetime

import openpyxl


# ── Parseo del Excel MGR ──────────────────────────────────────────────────────

def _parse_fecha(v) -> datetime:
    """Parsea datetime directo o string 'DD/MM/YYYY'/'YYYY-MM-DD'."""
    if isinstance(v, datetime):
        return v
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(v.strip(), fmt)
            except ValueError:
                pass
    return datetime(1900, 1, 1)


def _última_hoja_presupuesto(wb) -> Optional[str]:
    """Devuelve el nombre de la hoja de presupuesto con fecha más reciente.
    Busca por nombre 'Presupuesto' primero; si no hay, detecta hojas cuyo
    cell(1,5) contiene 'PRESUPUESTO'. Parsea fechas como datetime o string.
    Tie-break en empate de fechas: hoja con más líneas de contenido
    (evita elegir una hoja de 'lavandería' sobre la principal de 'cocina'
    en proyectos multi-zona con misma fecha)."""
    def _n_lineas(ws):
        n = 0
        for r in range(15, min(ws.max_row + 1, 50)):
            c4 = ws.cell(r, 4).value
            if c4 and str(c4).strip() not in ("", "#N/A", "0"):
                n += 1
        return n

    candidatos = []
    for name in wb.sheetnames:
        ws = wb[name]
        es_presup = "presupuesto" in name.lower()
        if not es_presup:
            cabecera = str(ws.cell(1, 5).value or "").upper()
            es_presup = "PRESUPUESTO" in cabecera
        if es_presup:
            candidatos.append((_parse_fecha(ws.cell(1, 4).value), _n_lineas(ws), name))
    if not candidatos:
        return None
    # Ordena por fecha desc, luego por # líneas desc (más contenido gana)
    candidatos.sort(key=lambda t: (t[0], t[1]), reverse=True)
    return candidatos[0][2]


def parse_excel_mgr(xlsx_path: Path) -> dict:
    """Lee el Excel y devuelve dict con campos normalizados para comparar."""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    hoja = _última_hoja_presupuesto(wb)
    if not hoja:
        return {"_error": "No se encontró hoja 'Presupuesto NNNN'"}
    ws = wb[hoja]

    result = {
        "_hoja":           hoja,
        "_fecha":          ws.cell(1, 4).value,
        "material":        None,
        "tablas":          None,
        "huecos":          {},      # {'placa': 1, 'fregadero_be': 1, 'grifo': 1, 'enchufe': 1, ...}
        "cantos_ml":       {},      # {'recto_pulido_agua': 4.92, 'bisel_pulido': 0, 'inglete': 2.3, ...}
        "chapeado_m2":     0.0,     # suma de "M2 COLOCACION CHAPEADO"
        "encimera_m2":     0.0,     # suma de "M2 COLOCACION ENCIMERA"
        "copete_ml":       0.0,     # "ML COPETE"
        "rodapie_ml":      0.0,     # "ML RODAPIE"
        "zocalo_ml":       0.0,
        "_raw_rows":       [],      # diagnóstico
    }

    # Palabras clave para identificar filas de MATERIAL (col 1) por heurística
    MATERIAL_KEYS = (
        "MM", "CM", "PULIDO", "MATE", "NATURAL", "APOMAZ", "DEKTON",
        "SILESTONE", "NEOLITH", "COMPAC", "COVERLAM", "LAMINAM", "GRANITO",
        "SILVESTRE", "PEDRAS", "MONDARIZ", "PERSIAN", "BELVEDERE", "LINEN",
        "BASIC", "ABSOLUTO", "IBERICO", "NERO", "BLANCO", "NEGRO", "GRIS",
    )

    for r in range(15, min(ws.max_row + 1, 120)):
        # Dos formatos conocidos:
        #   Nuevo (Cocimoble2026): col1=desc, col9=m²/cantid
        #   Viejo (ACyC/Cocimoble2025 antiguos): col1=None, col4=desc, col8=Unid
        # Algunos Excel tienen col1 con texto STALE del template (ej: 'SILVESTRE
        # LALIN' copiado de otro proyecto) mientras col4 tiene el dato real.
        # Regla: para filas de material (col5 ∈ {Tabla, Encimera, Chapeado, Copete, Rodapé}),
        # preferir col4 que es donde el sistema pone el material REAL. Para
        # filas de huecos/cantos, col1 suele estar vacía y col4 tiene la descripción.
        c1 = ws.cell(r, 1).value
        c4 = ws.cell(r, 4).value
        c5_up = str(ws.cell(r, 5).value or "").strip().upper()
        FILAS_MATERIAL = ("TABLA","ENCIMERA","CHAPEADO M2","COPETE ML","RODAPE > 9CM.","RODAPÉ > 9CM.")
        if any(c5_up.startswith(m) for m in FILAS_MATERIAL) and c4 not in (None, "", 0):
            desc = c4
        else:
            desc = c1 if (c1 not in (None, "", 0)) else (c4 or "")
        desc_u = str(desc).upper().strip()
        if not desc_u or desc_u.startswith("#N/A"):
            continue
        longo = ws.cell(r, 6).value
        ancho = ws.cell(r, 7).value
        c8    = ws.cell(r, 8).value
        c9    = ws.cell(r, 9).value
        # En formato nuevo la cantidad está en col 9 ("m²/Cantid.").
        # En formato viejo no hay col 9 y la cantidad está en col 8 ("Unid.").
        unid  = c8
        cant  = c9 if c9 is not None else c8

        # Material: primera fila con indicio de material en col 1 (heurística)
        if result["material"] is None:
            col5 = str(ws.cell(r, 5).value or "").strip().upper()
            parece_material = (
                col5 == "TABLA" or
                any(k in desc_u for k in ("MM ", "CM ", " MM", " CM")) or
                (any(k in desc_u for k in MATERIAL_KEYS)
                 and not desc_u.startswith(("UND ", "ML ", "M2 ", "M² ",
                                             "MEDICION", "DISEÑO", "COLOCAC",
                                             "CORTE", "UND COLOC")))
            )
            if parece_material:
                result["material"] = str(desc).strip()
                result["tablas"]   = cant if isinstance(cant, (int, float)) else unid
                result["_raw_rows"].append(("material", desc, cant))
                continue

        # Huecos: "UND HUECO PLACA", "UND HUECO BAJO ENCIMERA", "UND HUECO GRIFO", etc.
        if desc_u.startswith("UND HUECO PLACA"):
            result["huecos"]["placa"] = result["huecos"].get("placa", 0) + (cant or 1)
        elif "BAJO ENCIMERA" in desc_u and "HUECO" in desc_u:
            result["huecos"]["fregadero_be"] = result["huecos"].get("fregadero_be", 0) + (cant or 1)
        elif "SOBRE ENCIMERA" in desc_u and "HUECO" in desc_u:
            result["huecos"]["fregadero_se"] = result["huecos"].get("fregadero_se", 0) + (cant or 1)
        elif desc_u.startswith("UND HUECO GRIFO"):
            result["huecos"]["grifo"] = result["huecos"].get("grifo", 0) + (cant or 1)
        elif desc_u.startswith("UND HUECO ENCHUFE"):
            result["huecos"]["enchufe"] = result["huecos"].get("enchufe", 0) + (cant or 1)
        elif desc_u.startswith("UND HUECO FREGADERO") or desc_u.startswith("UND HUECO LAVAB"):
            # Fregadero/lavabo sin subtipo específico
            result["huecos"]["fregadero"] = result["huecos"].get("fregadero", 0) + (cant or 1)

        # Cantos — "ML CANTO RECTO PULIDO" (con o sin 'AGUA') mapea a recto_pulido_agua
        # (convención actual: ya no se diferencia NAT/PREF; solo apomazados usan seco)
        if desc_u.startswith("ML CANTO RECTO PULIDO"):
            result["cantos_ml"]["recto_pulido_agua"] = result["cantos_ml"].get("recto_pulido_agua", 0) + (cant or 0)
        elif desc_u.startswith("ML BISEL") or "BISEL" in desc_u:
            result["cantos_ml"]["bisel"] = result["cantos_ml"].get("bisel", 0) + (cant or 0)
        elif desc_u.startswith("ML INGLETADO") or "INGLET" in desc_u:
            result["cantos_ml"]["inglete"] = result["cantos_ml"].get("inglete", 0) + (cant or 0)
        elif "BOLEADO" in desc_u:
            result["cantos_ml"]["boleado"] = result["cantos_ml"].get("boleado", 0) + (cant or 0)
        elif "CANTO PILASTRA" in desc_u:
            result["cantos_ml"]["pilastra"] = result["cantos_ml"].get("pilastra", 0) + (cant or 0)

        # M² y ml de piezas (acepta variantes: "M2 COLOCACION ENCIMERA", "M² ...", "UND COLOCACION ENCIMERA...")
        if ("CHAPEADO" in desc_u and ("M2" in desc_u or "M²" in desc_u)):
            result["chapeado_m2"] += float(cant or 0)
        elif ("ENCIMERA" in desc_u and ("M2" in desc_u or "M²" in desc_u)
              and "BAJO" not in desc_u and "HUECO" not in desc_u):
            result["encimera_m2"] += float(cant or 0)
        elif ("COLOCACION ENCIMERA" in desc_u and "UND" in desc_u
              and "HUECO" not in desc_u and "BAJO" not in desc_u):
            # Las "UND COLOCACION ENCIMERA HASTA X m2" — X va en Longo o Cant
            m = re.search(r"(\d+[,\.]?\d*)\s*M", desc_u)
            if m:
                try:
                    result["encimera_m2"] += float(m.group(1).replace(",", "."))
                except ValueError:
                    pass
        elif desc_u.startswith("ML COPETE"):
            result["copete_ml"] += float(cant or 0)
        elif desc_u.startswith("ML RODAPIE") or desc_u.startswith("ML ROD"):
            result["rodapie_ml"] += float(cant or 0)
        elif desc_u.startswith("ML ZOCALO") or desc_u.startswith("ML ZOC"):
            result["zocalo_ml"] += float(cant or 0)

    return result


# ── Resumen del JSON extraído ─────────────────────────────────────────────────

def resumen_json(datos: dict) -> dict:
    """Agrega los datos del JSON en la misma estructura que parse_excel_mgr."""
    r = {
        "material":    None,
        "huecos":      {},
        "cantos_ml":   {},
        "chapeado_m2": 0.0,
        "encimera_m2": 0.0,
        "copete_ml":   0.0,
        "rodapie_ml":  0.0,
        "zocalo_ml":   0.0,
    }

    # Material: primero rol "encimera", luego CUALQUIER rol que contenga "encimera"
    # (encimera_opcion1, encimera_opcion2, etc.). Cuando hay opciones alternativas
    # no resueltas, guardamos TODAS las candidatas para que comparar() pueda
    # matchear contra el Excel si coincide con cualquiera.
    def _str_mat(m):
        partes = [m.get("marca",""), m.get("color",""),
                  f"{m.get('grosor_cm','')}cm" if m.get('grosor_cm') else ""]
        return " ".join(p for p in partes if p).strip()

    mats = datos.get("materiales") or []
    candidatas = []
    elegido = None
    for m in mats:
        if (m.get("rol") or "").lower() == "encimera":
            elegido = m; break
    if not elegido:
        for m in mats:
            if "encimera" in (m.get("rol") or "").lower():
                candidatas.append(m)
        if candidatas:
            elegido = candidatas[0]
    if not elegido and mats:
        elegido = mats[0]
    if elegido:
        r["material"] = _str_mat(elegido)
    # Lista completa de candidatas (incluye la elegida) para matching laxo
    if candidatas:
        r["material_candidates"] = [_str_mat(m) for m in candidatas]
    elif elegido:
        r["material_candidates"] = [_str_mat(elegido)]
    else:
        r["material_candidates"] = []

    # Huecos
    for h in datos.get("huecos") or []:
        tipo = (h.get("tipo") or "").lower()
        cant = h.get("cantidad") or 1
        subtipo = (h.get("subtipo") or "").lower()
        if tipo == "placa":
            r["huecos"]["placa"] = r["huecos"].get("placa", 0) + cant
        elif tipo == "fregadero":
            if "bajo" in subtipo:
                r["huecos"]["fregadero_be"] = r["huecos"].get("fregadero_be", 0) + cant
            elif "sobre" in subtipo:
                r["huecos"]["fregadero_se"] = r["huecos"].get("fregadero_se", 0) + cant
            else:
                r["huecos"]["fregadero"] = r["huecos"].get("fregadero", 0) + cant
        elif tipo == "grifo":
            r["huecos"]["grifo"] = r["huecos"].get("grifo", 0) + cant
        elif tipo == "enchufe":
            r["huecos"]["enchufe"] = r["huecos"].get("enchufe", 0) + cant

    # Cantos — "recto_pulido" y "recto_pulido_agua" se consideran el mismo bucket
    # salvo para materiales apomazados. Convención usuario 2026-04-22: pulido normal
    # siempre mapea a recto_pulido_agua en el comparativo contra Excel.
    for c in datos.get("cantos") or []:
        tipo = (c.get("tipo") or "").lower()
        ml = float(c.get("longitud_ml") or 0)
        if "recto" in tipo and "pulido" in tipo:
            r["cantos_ml"]["recto_pulido_agua"] = r["cantos_ml"].get("recto_pulido_agua", 0) + ml
        elif "bisel" in tipo:
            r["cantos_ml"]["bisel"] = r["cantos_ml"].get("bisel", 0) + ml
        elif "inglet" in tipo:  # cubre "inglete" e "ingletado"
            r["cantos_ml"]["inglete"] = r["cantos_ml"].get("inglete", 0) + ml
        elif "boleado" in tipo:
            r["cantos_ml"]["boleado"] = r["cantos_ml"].get("boleado", 0) + ml
        elif "pilastra" in tipo:
            r["cantos_ml"]["pilastra"] = r["cantos_ml"].get("pilastra", 0) + ml

    # m² y ml de piezas — SI hay opciones alternativas (material_rol con sufijo
    # _opcion1, _opcion2...), las piezas están duplicadas entre opciones con la
    # misma geometría. Contamos solo UNA opción para no duplicar áreas.
    import re as _re
    piezas = datos.get("piezas") or []
    sufijos = set()
    for p in piezas:
        mr = (p.get("material_rol") or "").lower()
        m = _re.search(r"_opcion(\d+)$", mr)
        if m:
            sufijos.add(m.group(1))
    hay_opciones = len(sufijos) > 1
    # Elegir la opción canónica (menor número)
    opcion_canon = min(sufijos, key=int) if sufijos else None

    for p in piezas:
        tipo = (p.get("tipo") or "").lower()
        mr = (p.get("material_rol") or "").lower()
        m = _re.search(r"_opcion(\d+)$", mr)
        # Si hay varias opciones, solo contar piezas de la canónica (o sin sufijo)
        if hay_opciones and m and m.group(1) != opcion_canon:
            continue
        largo = float(p.get("largo_mm") or 0) / 1000
        ancho = float(p.get("ancho_mm") or 0) / 1000
        alto  = float(p.get("altura_mm") or p.get("alto_mm") or 0) / 1000
        if tipo in ("encimera", "isla", "cascada"):
            if largo and ancho:
                r["encimera_m2"] += largo * ancho
        elif tipo in ("chapeado", "frontal", "pilastra", "costado"):
            if largo and alto:
                r["chapeado_m2"] += largo * alto
        elif tipo == "copete":
            r["copete_ml"] += largo
        elif tipo == "rodapie":
            r["rodapie_ml"] += largo
        elif tipo == "zocalo":
            r["zocalo_ml"] += largo

    return r


# ── Comparar ──────────────────────────────────────────────────────────────────

def _aprox_eq(a, b, tol_abs=0.5, tol_rel=0.05) -> bool:
    """Igualdad aproximada con tolerancia absoluta + relativa."""
    if a is None or b is None:
        return a == b
    try:
        a, b = float(a), float(b)
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()
    if a == 0 and b == 0:
        return True
    diff = abs(a - b)
    if diff <= tol_abs:
        return True
    return diff / max(abs(a), abs(b)) <= tol_rel


def comparar(excel_data: dict, json_data: dict) -> list[dict]:
    """Lista de filas (concepto, excel, json, match, delta)."""
    filas = []

    # Material: coincide si comparten ≥1 palabra NO-genérica (>3 letras y no
    # stopword como MM, CM, PULIDO, MATE, NATURAL...). El Excel en formato viejo
    # a veces solo trae color + acabado sin marca, por eso basta con 1 palabra
    # distintiva en común. Matchea contra CUALQUIERA de las candidatas si hay
    # opciones alternativas no resueltas.
    m_ex = (excel_data.get("material") or "").upper()
    candidatos_js = json_data.get("material_candidates") or [json_data.get("material") or ""]
    STOPWORDS = {"MM","CM","PULIDO","MATE","NATURAL","TEXTURIZADO","BRILLO",
                 "APOMAZADO","APOMAZ","SOFT","TOUCH","VELVET","ACABADO",
                 "NACIONAL","IMPORTACION","TABLA","ENCIMERA","CHAPEADO","PULIDA"}
    # Normalizar acentos sencillos para que "Nacré" matchee con "NACRE"
    import unicodedata
    def norm(s):
        return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn").upper()
    def palabras(s):
        import re as _re
        return set(w for w in _re.split(r"[ \-_,/\(\)]+", norm(s)) if len(w) > 3 and w not in STOPWORDS)
    match_mat = False
    mejor_js = candidatos_js[0] if candidatos_js else ""
    for cand in candidatos_js:
        if not m_ex or not cand:
            continue
        palabras_comunes = palabras(m_ex) & palabras(cand)
        if palabras_comunes:
            match_mat = True
            mejor_js = cand
            break
    etiqueta_js = json_data.get("material")
    if len(candidatos_js) > 1:
        etiqueta_js = " | ".join(candidatos_js) + (" (match)" if match_mat else " (none match)")
    filas.append({
        "concepto":   "material",
        "excel":      excel_data.get("material"),
        "json":       etiqueta_js,
        "match":      bool(match_mat),
        "severidad":  "alta" if not match_mat else "ok",
    })

    # Huecos: comparar fregadero como TOTAL (fregadero + _be + _se) para evitar
    # falsos positivos por categorización inconsistente. El Excel suele usar
    # "UND HUECO FREGADERO/LAVABO" sin subtipo; Claude sí emite subtipo.
    ex_h = dict(excel_data.get("huecos", {}))
    js_h = dict(json_data.get("huecos", {}))
    ex_fregadero_total = sum(ex_h.get(k, 0) for k in ("fregadero", "fregadero_be", "fregadero_se"))
    js_fregadero_total = sum(js_h.get(k, 0) for k in ("fregadero", "fregadero_be", "fregadero_se"))
    # Quitar las categorías de fregadero y comparar total aparte
    for k in ("fregadero", "fregadero_be", "fregadero_se"):
        ex_h.pop(k, None); js_h.pop(k, None)

    if ex_fregadero_total or js_fregadero_total:
        match = abs(ex_fregadero_total - js_fregadero_total) < 0.5
        filas.append({
            "concepto":  "hueco.fregadero (total)",
            "excel":     ex_fregadero_total,
            "json":      js_fregadero_total,
            "match":     match,
            "severidad": "alta" if not match else "ok",
        })

    # Resto de huecos (placa, grifo, enchufe, etc)
    claves_huecos = set(ex_h.keys()) | set(js_h.keys())
    for k in sorted(claves_huecos):
        v_ex = ex_h.get(k, 0)
        v_js = js_h.get(k, 0)
        match = (abs(v_ex - v_js) < 0.5)
        filas.append({
            "concepto":   f"hueco.{k}",
            "excel":      v_ex,
            "json":       v_js,
            "match":      match,
            "severidad":  "alta" if not match and v_ex != v_js else "ok",
        })

    # Cantos (ml)
    claves_cantos = set(excel_data.get("cantos_ml", {}).keys()) | set(json_data.get("cantos_ml", {}).keys())
    for k in sorted(claves_cantos):
        v_ex = excel_data.get("cantos_ml", {}).get(k, 0)
        v_js = json_data.get("cantos_ml", {}).get(k, 0)
        # Tolerancia canto: 1ml absoluto (feedback usuario 2026-04-22)
        match = _aprox_eq(v_ex, v_js, tol_abs=1.0, tol_rel=0.15)
        filas.append({
            "concepto":   f"canto.{k} (ml)",
            "excel":      round(v_ex, 2),
            "json":       round(v_js, 2),
            "match":      match,
            "severidad":  "media" if not match else "ok",
        })

    # m² y ml de piezas
    # NOTA: encimera_m2 NO se compara — el Excel cobra colocación en UND (no m²),
    # así que su valor sería siempre 0 y no es ground truth. Solo comparamos
    # chapeado_m2 y ml de tiras (copete/rodapié/zócalo) cuando ambos lados tienen valor.
    for concepto, a_ex, a_js, tol_abs, tol_rel, sev in [
        ("chapeado_m2", excel_data.get("chapeado_m2", 0), json_data.get("chapeado_m2", 0), 0.5, 0.15, "alta"),
        ("copete_ml",   excel_data.get("copete_ml",   0), json_data.get("copete_ml",   0), 0.5, 0.15, "media"),
        ("rodapie_ml",  excel_data.get("rodapie_ml",  0), json_data.get("rodapie_ml",  0), 0.5, 0.15, "media"),
        ("zocalo_ml",   excel_data.get("zocalo_ml",   0), json_data.get("zocalo_ml",   0), 0.5, 0.15, "baja"),
    ]:
        # Si una de las fuentes da 0 y la otra >0, puede ser por falta de línea en el
        # Excel (pricing UND vs m²), no un error real. Solo comparamos si Excel > 0.
        if a_ex == 0:
            continue
        match = _aprox_eq(a_ex, a_js, tol_abs=tol_abs, tol_rel=tol_rel)
        filas.append({
            "concepto":   concepto,
            "excel":      round(a_ex, 2),
            "json":       round(a_js, 2),
            "match":      match,
            "severidad":  sev if not match else "ok",
        })

    return filas


# ── Entrada/salida ────────────────────────────────────────────────────────────

def encontrar_json_y_excel(carpeta: Path) -> tuple[Optional[Path], Optional[Path]]:
    json_file = None
    excel_file = None
    for f in carpeta.glob("*_extraccion.json"):
        json_file = f; break
    xlsxs = list(carpeta.glob("*.xlsx"))
    if xlsxs:
        # Preferir el que no tenga 'archivo' en el nombre
        xlsxs.sort(key=lambda p: ("archivo" in p.name.lower(), -p.stat().st_mtime))
        excel_file = xlsxs[0]
    return json_file, excel_file


def verificar_proyecto(json_path: Path, excel_path: Path) -> dict:
    """Devuelve dict con job_id, filas de comparación y resumen."""
    datos = json.loads(json_path.read_text(encoding="utf-8"))
    excel_data = parse_excel_mgr(excel_path)
    json_data  = resumen_json(datos)
    filas = comparar(excel_data, json_data)

    return {
        "job_id":      datos.get("job_id", json_path.stem.replace("_extraccion","")),
        "excel":       excel_path.name,
        "hoja_excel":  excel_data.get("_hoja"),
        "fecha_excel": excel_data.get("_fecha"),
        "filas":       filas,
        "matches":     sum(1 for f in filas if f["match"]),
        "diffs":       sum(1 for f in filas if not f["match"]),
    }


def imprimir_proyecto(res: dict, verbose: bool = True):
    mm, dd = res["matches"], res["diffs"]
    pct = (mm / (mm + dd) * 100) if (mm + dd) else 0
    print(f"\n═══ {res['job_id']}  —  {mm}/{mm+dd} OK ({pct:.0f}%)  [{res['hoja_excel']}] ═══")
    if not verbose and dd == 0:
        return
    for f in res["filas"]:
        symbol = "✓" if f["match"] else "✗"
        sev = "" if f["match"] else f"  [{f['severidad']}]"
        print(f"  {symbol} {f['concepto']:<30} excel={f['excel']!s:>15}  json={f['json']!s:>15}{sev}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("ruta", nargs="?",
                        help="JSON de extracción, carpeta de proyecto, o carpeta raíz con --lote")
    parser.add_argument("--lote", action="store_true",
                        help="Modo batch: recorre todas las subcarpetas con extraccion.json")
    parser.add_argument("--csv", help="Guardar resumen en CSV (solo con --lote)")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Mostrar TODAS las filas aunque coincidan (default: solo diffs)")
    args = parser.parse_args()

    if not args.ruta:
        parser.print_help(); return 1

    ruta = Path(args.ruta)

    if args.lote:
        if not ruta.is_dir():
            print("✗ Con --lote, la ruta debe ser una carpeta."); return 1

        resultados = []
        for sub in sorted(ruta.iterdir()):
            if not sub.is_dir(): continue
            j, x = encontrar_json_y_excel(sub)
            if j and x:
                try:
                    res = verificar_proyecto(j, x)
                    resultados.append(res)
                    imprimir_proyecto(res, verbose=args.verbose)
                except Exception as e:
                    print(f"\n✗ Error en {sub.name}: {e}")

        # Resumen agregado
        print(f"\n{'═'*66}")
        print(f"  RESUMEN LOTE — {len(resultados)} proyectos")
        print(f"{'═'*66}")
        if resultados:
            total_m = sum(r["matches"] for r in resultados)
            total_d = sum(r["diffs"] for r in resultados)
            total = total_m + total_d
            pct = total_m / total * 100 if total else 0
            print(f"  Aciertos globales: {total_m}/{total}  ({pct:.1f}%)")

            # Discrepancias típicas (top 10)
            from collections import Counter
            diffs_por_concepto = Counter()
            for r in resultados:
                for f in r["filas"]:
                    if not f["match"]:
                        diffs_por_concepto[f["concepto"]] += 1
            if diffs_por_concepto:
                print(f"\n  Top discrepancias por concepto:")
                for c, n in diffs_por_concepto.most_common(10):
                    print(f"    {c:<30}  {n} proyectos")

        if args.csv and resultados:
            csv_path = Path(args.csv)
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["job_id", "concepto", "excel", "json", "match", "severidad"])
                for r in resultados:
                    for fila in r["filas"]:
                        w.writerow([r["job_id"], fila["concepto"], fila["excel"],
                                    fila["json"], fila["match"], fila["severidad"]])
            print(f"\n  CSV: {csv_path}")

        return 0

    # Modo individual
    if ruta.is_file():
        json_path = ruta
        carpeta = ruta.parent
    elif ruta.is_dir():
        carpeta = ruta
        json_path, _ = encontrar_json_y_excel(carpeta)
        if not json_path:
            print(f"✗ No hay *_extraccion.json en {carpeta}"); return 1
    else:
        print(f"✗ No existe: {ruta}"); return 1

    _, excel_path = encontrar_json_y_excel(carpeta)
    if not excel_path:
        print(f"✗ No hay xlsx en {carpeta}"); return 1

    res = verificar_proyecto(json_path, excel_path)
    imprimir_proyecto(res, verbose=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
